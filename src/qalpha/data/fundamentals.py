"""Screener.in fundamentals: parser + point-in-time store (Q_alpha.md §3.2, Phase 0b).

Screener's "Export to Excel" produces a consistent template; the **Data Sheet** tab holds dated
Profit & Loss and Balance Sheet blocks. We extract the raw fields needed for the Value, Quality and
Dividend factors and store them with an **effective date = report_date + reporting lag** so the
backtest only ever sees a year's financials after they were actually published (no look-ahead —
§14 criterion 2). Annual data with a lag is the standard, defensible Phase-0b approximation.

Banks/financials are detected (empty "Raw Material Cost" row) and flagged: EV/EBITDA and D/E don't
apply to them, so those are left NaN and the factor scorer renormalises over the available metrics.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd

DEFAULT_PARQUET = Path("data/fundamentals/fundamentals.parquet")
DEFAULT_LAG_DAYS = 90  # annual results are typically filed within ~3 months of FY-end

# 1-indexed rows in the Screener "Data Sheet" template (data starts at column 2).
_ROW = {
    "annual_date": 16,
    "other_income": 25,
    "depreciation": 26,
    "interest": 27,
    "pbt": 28,
    "net_profit": 30,
    "dividend": 31,
    "raw_material": 18,
    "equity_capital": 57,
    "reserves": 58,
    "borrowings": 59,
    "investments": 64,
    "cash": 69,
    "actual_shares": 70,
    "adjusted_shares_cr": 93,
}

# Columns persisted per (ticker, report_date) row.
FIELDS = [
    "net_profit",
    "dividend_amount",
    "equity_capital",
    "reserves",
    "borrowings",
    "cash",
    "investments",
    "shares_cr",
    "ebitda",
    "is_financial",
]


def _num(value: object) -> float | None:
    return float(value) if isinstance(value, int | float) else None


def parse_screener_xlsx(
    path: str | Path, ticker: str, lag_days: int = DEFAULT_LAG_DAYS
) -> pd.DataFrame:
    """Parse one Screener export into a tidy per-year DataFrame for ``ticker``."""
    ws = openpyxl.load_workbook(path, data_only=True)["Data Sheet"]

    label = ws.cell(_ROW["annual_date"], 1).value
    if label != "Report Date" or ws.cell(_ROW["net_profit"], 1).value != "Net profit":
        raise ValueError(f"unexpected Screener layout in {path} (template changed?)")

    def series(row: int) -> list[float | None]:
        return [_num(ws.cell(row, c).value) for c in range(2, ws.max_column + 1)]

    raw_dates = [ws.cell(_ROW["annual_date"], c).value for c in range(2, ws.max_column + 1)]
    dates = [pd.Timestamp(d).date() for d in raw_dates if d is not None]
    n = len(dates)

    cols = {name: series(row)[:n] for name, row in _ROW.items()}
    is_financial = all(v is None for v in cols["raw_material"])

    rows: list[dict[str, object]] = []
    for i, report_date in enumerate(dates):
        eq = cols["equity_capital"][i]
        res = cols["reserves"][i]
        # Share count: prefer actual count (row 70, absolute -> crore), fall back to adjusted (cr).
        actual = cols["actual_shares"][i]
        shares_cr = actual / 1e7 if actual else cols["adjusted_shares_cr"][i]
        ebitda: float | None = None
        if not is_financial:
            pbt, dep, intr = cols["pbt"][i], cols["depreciation"][i], cols["interest"][i]
            if pbt is not None and dep is not None and intr is not None:
                ebitda = pbt + dep + intr - (cols["other_income"][i] or 0.0)
        rows.append(
            {
                "ticker": ticker,
                "report_date": report_date,
                "effective_date": report_date + timedelta(days=lag_days),
                "net_profit": cols["net_profit"][i],
                "dividend_amount": cols["dividend"][i],
                "equity_capital": eq,
                "reserves": res,
                "borrowings": cols["borrowings"][i],
                "cash": cols["cash"][i],
                "investments": cols["investments"][i],
                "shares_cr": shares_cr,
                "ebitda": ebitda,
                "is_financial": is_financial,
            }
        )
    return pd.DataFrame(rows)


def _ticker_from_filename(path: Path) -> str:
    """Map a Screener filename to an NSE yfinance ticker (e.g. 'HDFC Bank.xlsx' -> 'HDFCBANK.NS')."""
    stem = path.stem.strip().upper().replace(" ", "")
    return stem if stem.endswith(".NS") else f"{stem}.NS"


def ingest_dir(
    raw_dir: str | Path = "data/fundamentals/raw",
    out_parquet: str | Path = DEFAULT_PARQUET,
    *,
    filename_to_ticker: dict[str, str] | None = None,
    lag_days: int = DEFAULT_LAG_DAYS,
) -> pd.DataFrame:
    """Parse every ``*.xlsx`` in ``raw_dir`` into one point-in-time fundamentals Parquet."""
    raw = Path(raw_dir)
    frames: list[pd.DataFrame] = []
    for xlsx in sorted(raw.glob("*.xlsx")):
        if filename_to_ticker and xlsx.name in filename_to_ticker:
            ticker = filename_to_ticker[xlsx.name]
        else:
            ticker = _ticker_from_filename(xlsx)
        frames.append(parse_screener_xlsx(xlsx, ticker, lag_days=lag_days))
    if not frames:
        raise ValueError(f"no .xlsx files found in {raw}")
    out = pd.concat(frames, ignore_index=True)
    out_path = Path(out_parquet)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out.to_parquet(out_path, index=False)
    return out


@dataclass
class FundamentalsStore:
    """Point-in-time view over parsed fundamentals: latest reported figures known by a given date."""

    data: pd.DataFrame

    @classmethod
    def from_parquet(cls, path: str | Path = DEFAULT_PARQUET) -> FundamentalsStore:
        df = pd.read_parquet(path)
        df["report_date"] = pd.to_datetime(df["report_date"]).dt.date
        df["effective_date"] = pd.to_datetime(df["effective_date"]).dt.date
        return cls(df)

    @property
    def tickers(self) -> list[str]:
        return sorted(self.data["ticker"].unique())

    def latest_for(self, tickers: list[str], as_of: date) -> pd.DataFrame:
        """Latest fundamentals per ticker whose ``effective_date <= as_of`` (indexed by ticker)."""
        df = self.data[(self.data["effective_date"] <= as_of) & (self.data["ticker"].isin(tickers))]
        if df.empty:
            return pd.DataFrame(columns=["ticker", *FIELDS]).set_index("ticker")
        latest = df.sort_values("report_date").groupby("ticker").tail(1).set_index("ticker")
        return latest[FIELDS]

    def dividend_years(self, ticker: str, as_of: date) -> int:
        """Consecutive most-recent years with a positive dividend, as known by ``as_of``."""
        df = self.data[
            (self.data["ticker"] == ticker) & (self.data["effective_date"] <= as_of)
        ].sort_values("report_date")
        count = 0
        for amount in reversed(df["dividend_amount"].tolist()):
            if amount is not None and amount > 0:
                count += 1
            else:
                break
        return count


def _main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Ingest Screener.in xlsx exports into Parquet.")
    parser.add_argument("--raw", default="data/fundamentals/raw")
    parser.add_argument("--out", default=str(DEFAULT_PARQUET))
    parser.add_argument("--lag-days", type=int, default=DEFAULT_LAG_DAYS)
    args = parser.parse_args()
    df = ingest_dir(args.raw, args.out, lag_days=args.lag_days)
    print(f"Parsed {df['ticker'].nunique()} tickers, {len(df)} year-rows -> {args.out}")
    print("Financials detected:", sorted(df[df["is_financial"]]["ticker"].unique()))


if __name__ == "__main__":
    _main()
