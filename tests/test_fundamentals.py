"""Fundamentals parser, point-in-time store, and Value/Quality/Dividend factor tests."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd

from qalpha.data.fundamentals import FIELDS, FundamentalsStore, parse_screener_xlsx
from qalpha.factors.fundamental import dividend, quality, value


def _write_template(path: Path, *, financial: bool) -> None:
    """Build a minimal Screener 'Data Sheet' workbook with two annual columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Sheet"
    rows = {
        16: ("Report Date", "2022-03-31", "2023-03-31"),
        18: ("Raw Material Cost", None if financial else 50, None if financial else 60),
        25: ("Other Income", 10, 12),
        26: ("Depreciation", 20, 22),
        27: ("Interest", 5, 6),
        28: ("Profit before tax", 100, 120),
        30: ("Net profit", 80, 95),
        31: ("Dividend Amount", 30, 35),
        57: ("Equity Share Capital", 10, 10),
        58: ("Reserves", 200, 240),
        59: ("Borrowings", 15, 18),
        64: ("Investments", 50, 55),
        69: ("Cash & Bank", 40, 45),
        70: ("No. of Equity Shares", 100_000_000, 100_000_000),
        93: ("Adjusted Equity Shares in Cr", 10.0, 10.0),
    }
    for r, (label, *vals) in rows.items():
        ws.cell(r, 1, label)
        for j, v in enumerate(vals, start=2):
            ws.cell(r, j, v)
    wb.save(path)


def test_parser_extracts_fields(tmp_path: Path) -> None:
    p = tmp_path / "ACME.xlsx"
    _write_template(p, financial=False)
    df = parse_screener_xlsx(p, "ACME.NS", lag_days=90)

    assert list(df["report_date"]) == [date(2022, 3, 31), date(2023, 3, 31)]
    assert df["effective_date"].iloc[0] == date(2022, 3, 31) + timedelta(days=90)
    row = df.iloc[1]
    assert row["net_profit"] == 95
    assert row["shares_cr"] == 10.0  # 100,000,000 / 1e7
    assert not row["is_financial"]
    # EBITDA = PBT + Dep + Interest - OtherIncome = 120 + 22 + 6 - 12 = 136
    assert row["ebitda"] == 136


def test_parser_detects_bank_and_nulls_ebitda(tmp_path: Path) -> None:
    p = tmp_path / "BANK.xlsx"
    _write_template(p, financial=True)
    df = parse_screener_xlsx(p, "BANK.NS")
    assert df["is_financial"].all()
    assert df["ebitda"].isna().all()


def _synthetic_store() -> FundamentalsStore:
    """Three IT stocks with different profitability + one with a dividend gap."""
    base = {
        "equity_capital": 10.0,
        "reserves": 200.0,
        "borrowings": 15.0,
        "cash": 40.0,
        "investments": 50.0,
        "shares_cr": 10.0,
        "ebitda": 100.0,
        "is_financial": False,
    }
    rows = []
    for ticker, npat, div in [("AAA.NS", 50, 10), ("BBB.NS", 100, 10), ("CCC.NS", 150, 0)]:
        for yr in (2021, 2022, 2023):
            rows.append(
                {
                    "ticker": ticker,
                    "report_date": date(yr, 3, 31),
                    "effective_date": date(yr, 6, 30),
                    "net_profit": float(npat),
                    "dividend_amount": float(div),
                    **base,
                }
            )
    return FundamentalsStore(pd.DataFrame(rows))


def test_store_point_in_time_no_lookahead() -> None:
    store = _synthetic_store()
    # As of 2022-05-01, FY2022 results (effective 2022-06-30) are NOT yet known -> FY2021 is latest.
    latest = store.latest_for(["AAA.NS"], date(2022, 5, 1))
    assert latest.loc["AAA.NS", "net_profit"] == 50.0
    assert list(latest.columns) == FIELDS


def test_dividend_years_counts_consecutive() -> None:
    store = _synthetic_store()
    as_of = date(2024, 1, 1)
    assert store.dividend_years("AAA.NS", as_of) == 3
    assert store.dividend_years("CCC.NS", as_of) == 0  # never paid


def test_value_quality_differentiate_within_sector() -> None:
    store = _synthetic_store()
    sectors = {"AAA.NS": "IT", "BBB.NS": "IT", "CCC.NS": "IT"}
    price = pd.Series({"AAA.NS": 100.0, "BBB.NS": 100.0, "CCC.NS": 100.0})
    as_of = date(2024, 1, 1)

    # Same price, higher net profit -> higher EPS -> lower P/E -> cheaper -> higher value score.
    val = value(store, price, sectors, as_of)
    assert val["CCC.NS"] > val["AAA.NS"]
    # Higher net profit on same net worth -> higher ROE -> higher quality.
    qual = quality(store, price, sectors, as_of)
    assert qual["CCC.NS"] > qual["AAA.NS"]


def test_dividend_factor_nan_for_non_payer() -> None:
    store = _synthetic_store()
    div = dividend(store, ["AAA.NS", "CCC.NS"], date(2024, 1, 1))
    assert div["AAA.NS"] == 3.0
    assert pd.isna(div["CCC.NS"])  # 0 years -> NaN so it drops out of the composite
